"""
Generic Excel reader — reads any Excel file, no hardcoded column names.

Returns raw rows as dicts {column_header: value}.
The LLM column mapper (audit/llm/column_mapper.py) handles semantic field mapping.
"""

from __future__ import annotations

import re
import openpyxl


def _clean(value) -> str:
    """Normalize any cell value to a clean string."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()



def get_sheet_names(file_path: str) -> list[str]:
    wb = openpyxl.load_workbook(file_path, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def read_excel_raw(
    file_path: str,
    sheet_name: str | None = None,
    max_rows: int | None = None,
    skip_empty: bool = True,
) -> tuple[list[str], list[dict[str, str]]]:
    """
    Read an Excel sheet and return (headers, rows).

    headers : list of column names from the first non-empty row
    rows    : list of dicts {header: cleaned_string_value}

    If sheet_name is None, reads the first sheet.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    headers: list[str] = []
    rows: list[dict[str, str]] = []

    for raw_row in ws.iter_rows(values_only=True):
        cleaned = [_clean(c) for c in raw_row]

        if not headers:
            if any(cleaned):
                headers = cleaned
            continue

        if skip_empty and not any(cleaned):
            continue

        if max_rows is not None and len(rows) >= max_rows:
            break

        rows.append(dict(zip(headers, cleaned)))

    wb.close()
    return headers, rows


def get_sample_rows(rows: list[dict[str, str]], n: int = 5) -> list[dict[str, str]]:
    """Return up to n rows with the most non-empty cells — no LLM involved."""
    scored = sorted(rows, key=lambda r: sum(1 for v in r.values() if v), reverse=True)
    return scored[:n]


def read_all_sheets_raw(
    file_path: str,
    max_rows: int | None = None,
) -> dict[str, tuple[list[str], list[dict[str, str]]]]:
    """
    Read every sheet in the workbook.
    Returns {sheet_name: (headers, rows)}.
    """
    result = {}
    for name in get_sheet_names(file_path):
        result[name] = read_excel_raw(file_path, sheet_name=name, max_rows=max_rows)
    return result
